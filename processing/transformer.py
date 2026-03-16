import pandas as pd
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

STAGE_ORDER = [
    ("Bulk Case Create API Invoke", ["CREATE_JOB"]),
    ("Web Job 1 (Load Data)", ["LOAD_TO_BLOB", "LOAD_TO_STG"]),
    ("Web Job 2 (Validate Data)", ["DATA_VALIDATE"]),
    ("Web Job 3 (Create Case & Report)", ["CREATE_CASE", "CREATE_REPORT"]),
]

RAW_COLS = [
    "JobStepId", "JobDefinitionStepsId", "JobId", "JobStepName",
    "Description", "ExecutionOrder", "JobDefinitionStepStatusId",
    "IsDeleted", "CreatedUser", "CreatedTimestamp",
    "UpdatedUser", "UpdatedTimestamp",
]


def _fmt_td(td):
    """Format a timedelta as hh:mm:ss.000"""
    if td is None:
        return ""
    total_seconds = td.total_seconds()
    sign = "-" if total_seconds < 0 else ""
    total_seconds = abs(total_seconds)
    hours, remainder = divmod(int(total_seconds), 3600)
    minutes, seconds = divmod(remainder, 60)
    millis = int((total_seconds - int(total_seconds)) * 1000)
    return f"{sign}{hours:02d}:{minutes:02d}:{seconds:02d}.{millis:03d}"


def _parse_ts(val):
    """Parse a timestamp value to datetime."""
    if isinstance(val, pd.Timestamp):
        return val.to_pydatetime()
    if isinstance(val, str):
        return pd.to_datetime(val).to_pydatetime()
    return val


def _assign_rows_to_stages(df):
    """Walk through rows sequentially, assigning to stages by JobStepName."""
    stages = []
    stage_idx = 0
    current_stage_name, current_step_names = STAGE_ORDER[stage_idx]
    current_rows = []

    for i, row in df.iterrows():
        step_name = row["JobStepName"]
        if step_name in current_step_names:
            current_rows.append(i)
        else:
            if current_rows:
                stages.append((current_stage_name, current_rows))
                current_rows = []
            stage_idx += 1
            while stage_idx < len(STAGE_ORDER):
                current_stage_name, current_step_names = STAGE_ORDER[stage_idx]
                if step_name in current_step_names:
                    current_rows.append(i)
                    break
                stage_idx += 1

    if current_rows:
        stages.append((current_stage_name, current_rows))

    return stages


def _compute_timings(df, stages):
    """Compute processing times and wait times between stages."""
    results = []
    prev_end_ts = None

    for s_idx, (stage_name, df_indices) in enumerate(stages):
        rows = df.iloc[df_indices]
        start_ts = _parse_ts(rows["UpdatedTimestamp"].iloc[0])
        end_ts = _parse_ts(rows["UpdatedTimestamp"].iloc[-1])

        # Wait time before this stage
        if prev_end_ts is not None and s_idx > 0:
            wait = start_ts - prev_end_ts
            results.append({
                "type": "wait",
                "label": f"Wait Time (before {stage_name})",
                "duration": wait,
            })

        # Processing time for this stage
        processing = end_ts - start_ts
        results.append({
            "type": "stage",
            "label": stage_name,
            "start": start_ts,
            "end": end_ts,
            "duration": processing,
            "rows": rows,
            "df_indices": df_indices,
        })

        prev_end_ts = end_ts

    # Final status row
    final_ts = _parse_ts(df["UpdatedTimestamp"].iloc[-1])
    first_ts = _parse_ts(df["UpdatedTimestamp"].iloc[0])
    total = final_ts - first_ts

    return results, total, first_ts, final_ts


def generate_report(df, case_count, template_path, output_path):
    stages = _assign_rows_to_stages(df)
    timings, total_time, first_ts, final_ts = _compute_timings(df, stages)

    wb = Workbook()
    ws = wb.active
    ws.title = "Performance Report"

    # Styles
    header_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11)
    time_font = Font(bold=True, size=11, color="2F5496")
    wait_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    stage_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    summary_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    row = 1

    # === SECTION 1: TIMING SUMMARY ===
    ws.cell(row=row, column=1, value="TIMING SUMMARY").font = Font(bold=True, size=14)
    row += 2

    summary_headers = ["Stage", "Start Time", "End Time", "Duration (hh:mm:ss.000)"]
    for ci, h in enumerate(summary_headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = header_font
        cell.border = thin_border
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    row += 1

    for timing in timings:
        if timing["type"] == "wait":
            ws.cell(row=row, column=1, value=timing["label"]).font = section_font
            for ci in range(1, 5):
                ws.cell(row=row, column=ci).fill = wait_fill
                ws.cell(row=row, column=ci).border = thin_border
            ws.cell(row=row, column=4, value=_fmt_td(timing["duration"])).font = time_font
            row += 1
        else:
            ws.cell(row=row, column=1, value=timing["label"]).font = section_font
            ws.cell(row=row, column=2, value=timing["start"].strftime("%Y-%m-%d %H:%M:%S"))
            ws.cell(row=row, column=3, value=timing["end"].strftime("%Y-%m-%d %H:%M:%S"))
            ws.cell(row=row, column=4, value=_fmt_td(timing["duration"])).font = time_font
            for ci in range(1, 5):
                ws.cell(row=row, column=ci).fill = stage_fill
                ws.cell(row=row, column=ci).border = thin_border
            row += 1

    # Total Processing Time
    row += 1
    for ci in range(1, 5):
        ws.cell(row=row, column=ci).fill = summary_fill
        ws.cell(row=row, column=ci).border = thin_border
    ws.cell(row=row, column=1, value="Total Processing Time").font = Font(bold=True, size=12)
    ws.cell(row=row, column=2, value=first_ts.strftime("%Y-%m-%d %H:%M:%S"))
    ws.cell(row=row, column=3, value=final_ts.strftime("%Y-%m-%d %H:%M:%S"))
    ws.cell(row=row, column=4, value=_fmt_td(total_time)).font = Font(bold=True, size=12, color="006100")
    row += 1

    # Time per Case
    if case_count and case_count > 0:
        time_per_case = total_time / case_count
        for ci in range(1, 5):
            ws.cell(row=row, column=ci).fill = summary_fill
            ws.cell(row=row, column=ci).border = thin_border
        ws.cell(row=row, column=1, value=f"Time per Case ({case_count} cases)").font = Font(bold=True, size=12)
        ws.cell(row=row, column=4, value=_fmt_td(time_per_case)).font = Font(bold=True, size=12, color="006100")
    row += 2

    # === SECTION 2: RAW JOB STEPS DATA ===
    ws.cell(row=row, column=1, value="RAW JOB STEPS DATA").font = Font(bold=True, size=14)
    row += 2

    display_cols = ["JobStepId", "JobStepName", "Description",
                    "JobDefinitionStepStatusId", "UpdatedTimestamp"]
    for ci, h in enumerate(display_cols, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    row += 1

    for _, data_row in df.iterrows():
        for ci, col_name in enumerate(display_cols, 1):
            val = data_row.get(col_name)
            if hasattr(val, "strftime"):
                val = val.strftime("%Y-%m-%d %H:%M:%S")
            cell = ws.cell(row=row, column=ci, value=val)
            cell.border = thin_border
        row += 1

    # Auto-fit column widths
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    wb.save(output_path)
    print(f"Report generated: {output_path}")
    return output_path
