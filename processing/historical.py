import re
import openpyxl
from datetime import timedelta


def _parse_duration(val):
    """Parse a duration string like '01:49:00' or '00:05:10.693000' to seconds."""
    if val is None or str(val).strip() in ("", "NA", "None"):
        return None
    s = str(val).strip()
    match = re.match(r"^(\d+):(\d+):(\d+)(?:\.(\d+))?$", s)
    if not match:
        return None
    h, m, sec = int(match.group(1)), int(match.group(2)), int(match.group(3))
    frac = match.group(4)
    micro = 0
    if frac:
        micro = int(frac.ljust(6, "0")[:6])
    return timedelta(hours=h, minutes=m, seconds=sec, microseconds=micro).total_seconds()


def load_run_log(filepath):
    """Load the RUN LOG sheet and return a list of execution records.

    Each execution has multiple 2K-file entries. We compute averages per execution.
    Returns list of dicts with keys:
        execution, date, test_type, test_name,
        avg_total, avg_wj1, avg_wj2, avg_wj3, avg_tpc
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["RUN LOG"]

    executions = []
    current_exec = None

    for row in range(4, ws.max_row + 1):
        sl_no = ws.cell(row=row, column=2).value
        date_val = ws.cell(row=row, column=3).value
        test_type = ws.cell(row=row, column=4).value
        test_name = ws.cell(row=row, column=5).value
        total_time = ws.cell(row=row, column=6).value
        wj1 = ws.cell(row=row, column=7).value
        wj2 = ws.cell(row=row, column=8).value
        wj3 = ws.cell(row=row, column=9).value
        tpc = ws.cell(row=row, column=11).value

        total_secs = _parse_duration(total_time)
        if total_secs is None:
            continue

        wj1_secs = _parse_duration(wj1)
        wj2_secs = _parse_duration(wj2)
        wj3_secs = _parse_duration(wj3)

        # Parse tpc (could be seconds as float or duration string)
        tpc_secs = None
        if tpc is not None:
            try:
                tpc_secs = float(tpc)
            except (ValueError, TypeError):
                tpc_secs = _parse_duration(tpc)

        if sl_no is not None and str(sl_no).strip():
            if current_exec is not None:
                executions.append(current_exec)
            date_str = ""
            if date_val:
                if hasattr(date_val, "strftime"):
                    date_str = date_val.strftime("%Y-%m-%d")
                else:
                    date_str = str(date_val).split(" ")[0]
            current_exec = {
                "execution": f"Exec #{sl_no}",
                "date": date_str,
                "test_type": str(test_type) if test_type else "",
                "test_name": str(test_name) if test_name else "",
                "totals": [],
                "wj1s": [],
                "wj2s": [],
                "wj3s": [],
                "tpcs": [],
            }

        if current_exec is not None:
            current_exec["totals"].append(total_secs)
            if wj1_secs is not None:
                current_exec["wj1s"].append(wj1_secs)
            if wj2_secs is not None:
                current_exec["wj2s"].append(wj2_secs)
            if wj3_secs is not None:
                current_exec["wj3s"].append(wj3_secs)
            if tpc_secs is not None:
                current_exec["tpcs"].append(tpc_secs)

    if current_exec is not None:
        executions.append(current_exec)

    # Compute averages
    results = []
    for ex in executions:
        avg = lambda lst: sum(lst) / len(lst) if lst else 0
        results.append({
            "execution": ex["execution"],
            "date": ex["date"],
            "test_type": ex["test_type"],
            "test_name": ex["test_name"],
            "file_count": len(ex["totals"]),
            "avg_total": avg(ex["totals"]),
            "avg_wj1": avg(ex["wj1s"]),
            "avg_wj2": avg(ex["wj2s"]),
            "avg_wj3": avg(ex["wj3s"]),
            "avg_tpc": avg(ex["tpcs"]),
        })

    return results
